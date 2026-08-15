"""
Vercel serverless endpoint: POST /api/upload-sales
Accepts a sales-per-day file (CSV or Excel) uploaded from the dashboard and
writes it into the daily_sales table. Handles the Bagisto "Sales Per Day"
export format (semicolon-delimited, DD-MM-YYYY dates, Net Total = revenue)
as well as the already-converted format (comma, YYYY-MM-DD, revenue column).

Returns JSON: {"ok": true, "imported": N, "range": "start..end"} or {"error": ...}
"""
import os
import io
import csv
import json
from datetime import datetime
from http.server import BaseHTTPRequestHandler

import psycopg2


def _get_conn():
    url = os.environ["DATABASE_URL"]
    if "sslmode=" not in url:
        url += ("&" if "?" in url else "?") + "sslmode=require"
    return psycopg2.connect(url)


def _parse_date(s):
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _num(s):
    s = str(s or "").replace(",", "").replace("$", "").replace("€", "").replace("£", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _find_col(low_headers, candidates):
    """Return the index of the first header matching a candidate term.
    Matches exact, startswith, or contains — so 'net total (excl. tax)'
    matches the candidate 'net total'. Prefers the earliest, most exact match."""
    # exact first
    for cand in candidates:
        if cand in low_headers:
            return low_headers.index(cand)
    # then startswith
    for cand in candidates:
        for i, h in enumerate(low_headers):
            if h.startswith(cand):
                return i
    # then contains
    for cand in candidates:
        for i, h in enumerate(low_headers):
            if cand in h:
                return i
    return None


def _rows_from_csv(text):
    """Yield (date, revenue, orders) from CSV text in either Bagisto or converted format."""
    lines = text.splitlines()
    # strip Bagisto "sep=;" preamble
    if lines and lines[0].lower().startswith("sep="):
        lines = lines[1:]
    if not lines:
        return
    # detect delimiter from header
    header = lines[0]
    delim = ";" if header.count(";") >= header.count(",") else ","
    reader = csv.DictReader(lines, delimiter=delim)
    reader.fieldnames = [(h or "").strip() for h in (reader.fieldnames or [])]
    low_headers = [f.lower() for f in reader.fieldnames]
    # map columns flexibly (handles 'Net Total (excl. Tax)' etc.)
    di = _find_col(low_headers, ("date",))
    ri = _find_col(low_headers, ("net total", "revenue", "net"))
    oi = _find_col(low_headers, ("sales", "orders"))
    date_col = reader.fieldnames[di] if di is not None else None
    rev_col = reader.fieldnames[ri] if ri is not None else None
    ord_col = reader.fieldnames[oi] if oi is not None else None
    for row in reader:
        row = {(k or "").strip(): v for k, v in row.items()}
        raw_date = (row.get(date_col, "") or "").strip() if date_col else ""
        if raw_date.lower() in ("totals", "total", ""):
            continue
        d = _parse_date(raw_date) if date_col else None
        if not d:
            continue
        rev = _num(row.get(rev_col, "")) if rev_col else None
        orders = _num(row.get(ord_col, "")) if ord_col else None
        yield d, rev, int(orders) if orders is not None else None


def _rows_from_xlsx(data):
    """Yield (date, revenue, orders) from an .xlsx byte payload using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = None
    idx_date = idx_rev = idx_ord = None
    for r in rows:
        if r is None:
            continue
        cells = [str(c).strip() if c is not None else "" for c in r]
        low = [c.lower() for c in cells]
        if header is None:
            if "date" in low:
                header = low
                idx_date = low.index("date")
                idx_rev = _find_col(low, ("net total", "revenue", "net"))
                idx_ord = _find_col(low, ("sales", "orders"))
            continue
        if idx_date is None or idx_date >= len(r):
            continue
        # skip summary/total rows
        first = str(r[idx_date]).strip().lower() if r[idx_date] is not None else ""
        if first in ("totals", "total", ""):
            continue
        d = _parse_date(str(r[idx_date])) if r[idx_date] is not None else None
        if not d:
            continue
        rev = _num(r[idx_rev]) if (idx_rev is not None and idx_rev < len(r)) else None
        orders = _num(r[idx_ord]) if (idx_ord is not None and idx_ord < len(r)) else None
        yield d, rev, int(orders) if orders is not None else None


def _extract_file(content_type, body):
    """Pull the uploaded file bytes + filename out of a multipart/form-data body.
    Uses the stdlib email parser, which handles the exact byte framing browsers
    send (boundary matching, CRLFs, binary payloads) far more reliably than a
    manual split."""
    if "boundary=" not in (content_type or ""):
        return None, None
    from email.parser import BytesParser
    from email.policy import default as default_policy
    # reconstruct a minimal MIME message: headers + blank line + body
    header = ("Content-Type: " + content_type + "\r\n\r\n").encode()
    msg = BytesParser(policy=default_policy).parsebytes(header + body)
    for part in msg.iter_parts():
        cd = part.get("Content-Disposition", "")
        if "filename" in cd:
            fname = part.get_filename() or ""
            payload = part.get_payload(decode=True) or b""
            return fname, payload
    return None, None


class handler(BaseHTTPRequestHandler):
    def _send(self, code, obj):
        body = json.dumps(obj).encode()
        self.send_response(code)
        self.send_header("Content-Type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        # lets you sanity-check the endpoint in a browser; confirms it deployed
        # and can reach the database
        try:
            conn = _get_conn()
            cur = conn.cursor()
            cur.execute("SELECT count(*), min(sale_date), max(sale_date) FROM daily_sales")
            n, lo, hi = cur.fetchone()
            cur.close(); conn.close()
            self._send(200, {"ok": True, "endpoint": "upload-sales",
                             "message": "POST a CSV/Excel file here to import sales.",
                             "current_rows": n,
                             "range": f"{lo} .. {hi}" if lo else None})
        except Exception as e:
            self._send(500, {"error": f"DB check failed: {e}"})

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            ctype = self.headers.get("Content-Type", "")
            fname, payload = _extract_file(ctype, body)
            if not payload:
                return self._send(400, {"error": "No file found in upload."})

            name = (fname or "").lower()
            if name.endswith(".xlsx") or name.endswith(".xlsm"):
                rows = list(_rows_from_xlsx(payload))
            else:
                # assume CSV/text
                text = payload.decode("utf-8-sig", "ignore")
                rows = list(_rows_from_csv(text))

            if not rows:
                return self._send(400, {"error": "No valid sales rows found. Expected a Date column and a Net Total/Revenue column."})

            # safety guard: if NONE of the rows have revenue, the revenue column
            # almost certainly didn't match — refuse rather than wipe existing data
            rows_with_rev = [r for r in rows if r[1] is not None]
            if not rows_with_rev:
                return self._send(400, {"error": "Found dates but no revenue values — the revenue column wasn't recognized (expected something like 'Net Total'). Nothing was changed."})

            conn = _get_conn()
            cur = conn.cursor()
            n = 0
            for d, rev, orders in rows:
                cur.execute("""
                    INSERT INTO daily_sales (sale_date, revenue, orders, updated_at)
                    VALUES (%s, %s, %s, now())
                    ON CONFLICT (sale_date) DO UPDATE SET
                      revenue = COALESCE(EXCLUDED.revenue, daily_sales.revenue),
                      orders  = COALESCE(EXCLUDED.orders, daily_sales.orders),
                      updated_at = now()
                """, (d.isoformat(), rev, orders))
                n += 1
            conn.commit()
            # record the import timestamp
            try:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS data_updates (
                      kind TEXT PRIMARY KEY, updated_at TIMESTAMPTZ NOT NULL,
                      source TEXT, detail TEXT)
                """)
                cur.execute("""
                    INSERT INTO data_updates (kind, updated_at, source, detail)
                    VALUES ('sales', now(), 'import', %s)
                    ON CONFLICT (kind) DO UPDATE SET
                      updated_at=now(), source='import', detail=EXCLUDED.detail
                """, (f"{n} days imported",))
                conn.commit()
            except Exception:
                conn.rollback()
            cur.close()
            conn.close()

            ds = sorted(r[0] for r in rows)
            return self._send(200, {
                "ok": True, "imported": n,
                "range": f"{ds[0].isoformat()} .. {ds[-1].isoformat()}",
            })
        except KeyError:
            return self._send(500, {"error": "DATABASE_URL not configured on the server."})
        except Exception as e:
            return self._send(500, {"error": f"Import failed: {e}"})
